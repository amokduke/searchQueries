import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


INPUT_FILE = Path("CAREkakis Volunteers Performance Matrix_R.xlsx")
SEARCH_CONSTITUENCY = "Tampines West"
OUTPUT_FILE = Path("Tampines West_events_and_volunteers.xlsx")


EVENT_COLUMNS = [
    "Date",
    "Day",
    "Time",
    "Venue",
    "Program Stage",
    "Program Name",
    "Remarks",
    "Staff",
    "Trainer",
    "Registered",
    "GRL Attended",
    "Public Attended CA",
    "Total Attendance",
    "Source Sheet",
    "Original Header",
]

VOLUNTEER_COLUMNS = [
    "Name",
    "Email",
    "Handphone",
    "Constituency",
    "Source Sheet",
    "Recorded Event Count",
    "Recorded Events",
]

APPEARANCE_COLUMNS = [
    "Name",
    "Email",
    "Handphone",
    "Constituency",
    "Date",
    "Day",
    "Time",
    "Venue",
    "Program Stage",
    "Program Name",
    "Remarks",
    "Source Sheet",
    "Original Header",
]


VENUE_MAP = {
    "BL": "Boon Lay",
    "KG": "Kampong Glam",
    "KPG": "Kampong Glam",
    "BBE": "Bukit Batok East",
    "TB": "Tampines",
    "TW": "Tampines West",
    "VAD": "Varies / To confirm",
}


def normalise_text(value):
    if value is None:
        return ""
    return str(value).strip()


def canonicalise_header(text):
    text = normalise_text(text).lower()
    text = re.sub(r"[\n\r\t]+", " ", text)
    text = re.sub(r"[^a-z0-9+ ]+", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def is_target_header_row(row_values):
    row_norm = [canonicalise_header(v) for v in row_values]
    return "sn" in row_norm and "name" in row_norm and "constituency" in row_norm


def find_header_row(ws, max_scan_rows=10):
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 25) + 1)]
        if is_target_header_row(values):
            return r
    return None


def is_attendance_mark(value):
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip()
    if text == "":
        return False
    if text in {"-", "0", "N", "No", "NO", "n", "no"}:
        return False
    return True


def header_looks_like_event(header_text):
    if not header_text:
        return False

    text = str(header_text).strip()

    ignore_phrases = [
        "total cc",
        "total ck",
        "total csg",
        "total cc + csg",
        "cumulative total",
        "no. of touchpoints",
        "signed up as volunteer",
        "training completed",
        "statement of attainment",
        "certification",
        "t-shirt",
        "email",
        "e-mail",
        "handphone",
        "phone",
        "mobile",
        "constituency",
        "name",
        "s/n",
        "sn",
        "checklist",
        "javin update",
        "updated as of",
        "legend",
    ]

    lowered = text.lower()
    if any(phrase in lowered for phrase in ignore_phrases):
        return False

    date_patterns = [
        r"\b\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4}\b",
        r"\b\d{1,2}\s+[A-Za-z]{3,9}\b",
        r"\b\d{1,2}\s+[A-Za-z]{3,9}\s*:",
        r"\bpre-?ns\b",
        r"\bpurple parade\b",
    ]

    return any(re.search(p, text, flags=re.IGNORECASE) for p in date_patterns)


def parse_event_header(header_text, source_sheet):
    original = normalise_text(header_text).replace("\n", " ").strip()
    compact = re.sub(r"\s+", " ", original)

    date_value = None
    day_value = ""
    time_value = ""
    venue_value = ""
    program_stage = ""
    program_name = compact
    remarks = ""

    full_date_patterns = [
        r"(?P<date>\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4})",
        r"(?P<date>\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2})",
    ]

    for pattern in full_date_patterns:
        m = re.search(pattern, compact, flags=re.IGNORECASE)
        if m:
            raw_date = m.group("date")
            for fmt in ("%d %b %Y", "%d %B %Y", "%d %b %y", "%d %B %y"):
                try:
                    date_value = datetime.strptime(raw_date, fmt).date()
                    break
                except ValueError:
                    continue
            if date_value:
                break

    if date_value is None:
        inferred_year = None
        year_match = re.search(r"(20\d{2}|24|23)", source_sheet)
        if year_match:
            y = year_match.group(1)
            inferred_year = int("20" + y) if len(y) == 2 else int(y)

        partial_date_match = re.search(
            r"(?P<date>\d{1,2}\s+[A-Za-z]{3,9})(?:\s*:|\b)",
            compact,
            flags=re.IGNORECASE,
        )
        if partial_date_match and inferred_year:
            raw_date = f"{partial_date_match.group('date')} {inferred_year}"
            for fmt in ("%d %b %Y", "%d %B %Y"):
                try:
                    date_value = datetime.strptime(raw_date, fmt).date()
                    break
                except ValueError:
                    continue

    if date_value:
        day_value = date_value.strftime("%A")

    venue_match = re.search(
        r"^\s*\d{1,2}\s+[A-Za-z]{3,9}(?:\s+\d{4})?\s*:\s*([A-Za-z]{2,4})\s*:",
        compact,
        flags=re.IGNORECASE,
    )
    if venue_match:
        venue_code = venue_match.group(1).upper()
        venue_value = VENUE_MAP.get(venue_code, venue_code)

    program_name = re.sub(
        r"^\s*\d{1,2}\s+[A-Za-z]{3,9}(?:\s+\d{4})?\s*[:\-]?\s*",
        "",
        compact,
        flags=re.IGNORECASE,
    ).strip()

    program_name = re.sub(r"^[A-Za-z]{2,4}\s*:\s*", "", program_name).strip()

    if "pre-ns forum" in compact.lower():
        program_name = compact

    return {
        "Date": date_value,
        "Day": day_value,
        "Time": time_value,
        "Venue": venue_value,
        "Program Stage": program_stage,
        "Program Name": program_name,
        "Remarks": remarks,
        "Staff": "",
        "Trainer": "",
        "Registered": 0,
        "GRL Attended": "",
        "Public Attended CA": "",
        "Total Attendance": 0,
        "Source Sheet": source_sheet,
        "Original Header": compact,
    }


def find_column(headers, candidates):
    """
    headers: dict[col_index] = original header text
    candidates: list of possible canonical header names
    """
    candidate_set = {canonicalise_header(x) for x in candidates}

    for c, h in headers.items():
        if canonicalise_header(h) in candidate_set:
            return c

    for c, h in headers.items():
        canon = canonicalise_header(h)
        for candidate in candidate_set:
            if candidate in canon or canon in candidate:
                return c

    return None


def extract_events_and_volunteers(input_file, constituency):
    wb = load_workbook(input_file, data_only=True)

    event_summary_rows = []
    volunteer_rows = []
    appearance_rows = []

    seen_volunteers = set()
    seen_event_summary = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = find_header_row(ws)
        if header_row is None:
            continue

        headers = {
            c: normalise_text(ws.cell(header_row, c).value)
            for c in range(1, ws.max_column + 1)
        }

        constituency_col = find_column(headers, ["Constituency"])
        name_col = find_column(headers, ["Name"])
        email_col = find_column(headers, ["Email", "E-mail"])
        handphone_col = find_column(headers, ["Handphone", "Phone", "Mobile"])

        if constituency_col is None:
            continue

        event_cols = [c for c, h in headers.items() if header_looks_like_event(h)]
        if not event_cols:
            continue

        matching_rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            cell_value = normalise_text(ws.cell(r, constituency_col).value)
            if cell_value.lower() == constituency.strip().lower():
                matching_rows.append(r)

        if not matching_rows:
            continue

        event_attendance_counter = {c: 0 for c in event_cols}

        for r in matching_rows:
            volunteer_name = normalise_text(ws.cell(r, name_col).value) if name_col else ""
            volunteer_email = normalise_text(ws.cell(r, email_col).value) if email_col else ""
            volunteer_handphone = normalise_text(ws.cell(r, handphone_col).value) if handphone_col else ""
            volunteer_constituency = normalise_text(ws.cell(r, constituency_col).value)

            volunteer_event_list = []

            for c in event_cols:
                value = ws.cell(r, c).value
                if is_attendance_mark(value):
                    event_attendance_counter[c] += 1

                    event_info = parse_event_header(headers[c], sheet_name)
                    volunteer_event_list.append(event_info["Program Name"])

                    appearance_rows.append({
                        "Name": volunteer_name,
                        "Email": volunteer_email,
                        "Handphone": volunteer_handphone,
                        "Constituency": volunteer_constituency,
                        "Date": event_info["Date"],
                        "Day": event_info["Day"],
                        "Time": event_info["Time"],
                        "Venue": event_info["Venue"],
                        "Program Stage": event_info["Program Stage"],
                        "Program Name": event_info["Program Name"],
                        "Remarks": event_info["Remarks"],
                        "Source Sheet": sheet_name,
                        "Original Header": event_info["Original Header"],
                    })

            volunteer_key = (
                volunteer_name.lower(),
                volunteer_email.lower(),
                volunteer_handphone.lower(),
                volunteer_constituency.lower(),
            )

            if volunteer_key not in seen_volunteers:
                seen_volunteers.add(volunteer_key)
                volunteer_rows.append({
                    "Name": volunteer_name,
                    "Email": volunteer_email,
                    "Handphone": volunteer_handphone,
                    "Constituency": volunteer_constituency,
                    "Source Sheet": sheet_name,
                    "Recorded Event Count": len(volunteer_event_list),
                    "Recorded Events": "; ".join(sorted(set(volunteer_event_list))),
                })
            else:
                # Update existing volunteer if same person appears across sheets
                for row in volunteer_rows:
                    existing_key = (
                        row["Name"].lower(),
                        row["Email"].lower(),
                        row["Handphone"].lower(),
                        row["Constituency"].lower(),
                    )
                    if existing_key == volunteer_key:
                        existing_events = set(
                            x.strip() for x in row["Recorded Events"].split(";") if x.strip()
                        )
                        new_events = set(volunteer_event_list)
                        merged = sorted(existing_events | new_events)
                        row["Recorded Events"] = "; ".join(merged)
                        row["Recorded Event Count"] = len(merged)
                        break

        for c in event_cols:
            attended_count = event_attendance_counter[c]
            if attended_count > 0:
                event_row = parse_event_header(headers[c], sheet_name)
                event_key = (sheet_name, event_row["Original Header"])

                if event_key not in seen_event_summary:
                    seen_event_summary.add(event_key)
                    event_row["Registered"] = attended_count
                    event_row["Total Attendance"] = attended_count
                    event_summary_rows.append(event_row)

    events_df = pd.DataFrame(event_summary_rows, columns=EVENT_COLUMNS)
    volunteers_df = pd.DataFrame(volunteer_rows, columns=VOLUNTEER_COLUMNS)
    appearances_df = pd.DataFrame(appearance_rows, columns=APPEARANCE_COLUMNS)

    return events_df, volunteers_df, appearances_df


def main():
    events_df, volunteers_df, appearances_df = extract_events_and_volunteers(
        INPUT_FILE,
        SEARCH_CONSTITUENCY,
    )

    if events_df.empty and volunteers_df.empty:
        print(f'No records found for constituency: "{SEARCH_CONSTITUENCY}"')
        print("This may mean either:")
        print("1. there are no matching volunteer rows for that constituency in this workbook, or")
        print("2. the constituency label is different from what you searched.")
        return

    if not events_df.empty:
        events_df["_sort_date"] = pd.to_datetime(events_df["Date"], errors="coerce")
        events_df = events_df.sort_values(
            by=["_sort_date", "Program Name"],
            na_position="last"
        ).drop(columns=["_sort_date"])

    if not volunteers_df.empty:
        volunteers_df = volunteers_df.sort_values(
            by=["Recorded Event Count", "Name"],
            ascending=[False, True],
            na_position="last"
        )

    if not appearances_df.empty:
        appearances_df["_sort_date"] = pd.to_datetime(appearances_df["Date"], errors="coerce")
        appearances_df = appearances_df.sort_values(
            by=["Name", "_sort_date", "Program Name"],
            na_position="last"
        ).drop(columns=["_sort_date"])

    print("\n=== EVENTS SUMMARY ===")
    if events_df.empty:
        print("No events found.")
    else:
        print(events_df.to_string(index=False))

    print("\n=== VOLUNTEER BASE ===")
    if volunteers_df.empty:
        print("No volunteers found.")
    else:
        print(volunteers_df.to_string(index=False))

    print("\n=== VOLUNTEER EVENT APPEARANCES ===")
    if appearances_df.empty:
        print("No volunteer appearances found.")
    else:
        print(appearances_df.to_string(index=False))

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        if not events_df.empty:
            events_df.to_excel(writer, index=False, sheet_name="Events Summary")
        if not volunteers_df.empty:
            volunteers_df.to_excel(writer, index=False, sheet_name="Volunteer Base")
        if not appearances_df.empty:
            appearances_df.to_excel(writer, index=False, sheet_name="Volunteer Event Appearances")

    print(f"\nSaved to: {OUTPUT_FILE.resolve()}")


if __name__ == "__main__":
    main()